import os
import sys
import time
import json
import zipfile
import re
import datetime
import threading
from urllib.parse import parse_qs, urlparse
from socketserver import ThreadingMixIn
from http.server import HTTPServer, BaseHTTPRequestHandler
import xml.etree.ElementTree as ET

# Paths
BASE_ONEDRIVE_DIR = '/Users/netocerasi/Library/CloudStorage/OneDrive-LeoMadeiras/1. PLANILHA DE DADOS COLABORADORES PARA ACOES'
PARTNERS_DIR = os.path.join(BASE_ONEDRIVE_DIR, '01. Relacao de funcionarios parceiros')
CONSOLIDATED_DIR = os.path.join(BASE_ONEDRIVE_DIR, '02. Consolidado')

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_FILE = os.path.join(CURRENT_DIR, 'data_cache.json')
CONSOLIDATED_CSV = os.path.join(CURRENT_DIR, 'CONSOLIDADO_AUTOMATICO.csv')
CONSOLIDATED_XLSX = os.path.join(CURRENT_DIR, 'CONSOLIDADO_PARCEIROS.xlsx')
USERS_FILE = os.path.join(CURRENT_DIR, 'users.json')
CUSTOM_EMPLOYEES_FILE = os.path.join(CURRENT_DIR, 'custom_employees.json')

# Master Dictionary of Store Regions requested by User
STORE_REGIONS_MAP = {
    "2050": "BRASÍLIA",
    "2051": "CAMPINAS",
    "2052": "VALPARAÍSO DE GOIÁS",
    "2061": "GOIÂNIA",
    "2075": "GOIÂNIA",
    "2100": "SÃO LUÍS",
    "2101": "SÃO LUÍS COHAMA",
    "2102": "SANTA INÊS",
    "2103": "IMPERATRIZ",
    "2125": "CUIABÁ",
    "2126": "VÁRZEA GRANDE",
    "2127": "VÁRZEA GRANDE",
    "2400": "RIBEIRÃO PRETO",
    "2425": "CAMPINAS",
    "2426": "VÁRZEA GRANDE",
    "2427": "CUIABÁ",
    "2429": "CUIABÁ",
    "2475": "CURITIBA",
    "2550": "TERESINA",
    "2551": "TERESINA",
    "2552": "TERESINA",
    "2575": "UBERLÂNDIA",
    "2700": "PETROLINA",
    "2775": "CARUARU",
    "2800": "RECIFE",
    "2805": "RECIFE",
    "2808": "CAMPINA GRANDE",
    "2809": "CABEDELO",
    "2810 CD": "JOÃO PESSOA (CD)",
    "2825": "MACEIÓ",
    "2900": "BELO HORIZONTE",
    "2950": "CONTAGEM",
    "2952": "VENDA NOVA",
    "2955": "IPATINGA",
    "2956": "JUIZ DE FORA",
    "2962": "UBERABA",
    "2963": "SETE LAGOAS",
    "2964": "CASTELO"
}

PREDEFINED_STORES = [
    {"loja_num": num, "nome_loja": region} for num, region in STORE_REGIONS_MAP.items()
]

# Global State
global_state = {
    'last_updated': 'Aguardando inicialização...',
    'today_date': datetime.date.today().strftime('%d/%m/%Y'),
    'current_year': datetime.date.today().year,
    'stores_count': len(PREDEFINED_STORES),
    'total_records': 0,
    'active_records': 0,
    'terminated_records': 0,
    'mothers_fathers_count': 0,
    'total_children_count': 0,
    'children_by_age': {},
    'admission_years': [],
    'pcd_count': 0,
    'stores': PREDEFINED_STORES,
    'employees': [],
    'file_mtimes': {},
    'subscribers': []
}

global_lock = threading.Lock()

# Initial Default Users
DEFAULT_USERS = [
    {
        'id': 'usr_admin',
        'nome': 'Neto Cerasi (Admin)',
        'email': 'admin@leomadeiras.com.br',
        'senha': 'admin',
        'role': 'ADMIN',
        'stores': ['ALL']
    },
    {
        'id': 'usr_gerente_2050_2051',
        'nome': 'Gerente Região Brasília / Campinas',
        'email': 'gerente2050@leomadeiras.com.br',
        'senha': '123',
        'role': 'USER',
        'stores': ['2050', '2051']
    },
    {
        'id': 'usr_gerente_2050',
        'nome': 'Gerente Loja 2050 Brasília',
        'email': 'brasilia@leomadeiras.com.br',
        'senha': '123',
        'role': 'USER',
        'stores': ['2050']
    }
]

def load_users():
    if not os.path.exists(USERS_FILE):
        save_users(DEFAULT_USERS)
        return DEFAULT_USERS
    try:
        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return DEFAULT_USERS

def save_users(users_list):
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(users_list, f, ensure_ascii=False, indent=2)

def load_custom_employees():
    if not os.path.exists(CUSTOM_EMPLOYEES_FILE):
        return []
    try:
        with open(CUSTOM_EMPLOYEES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return []

def save_custom_employees(emp_list):
    with open(CUSTOM_EMPLOYEES_FILE, 'w', encoding='utf-8') as f:
        json.dump(emp_list, f, ensure_ascii=False, indent=2)

# Column Headers Mapping (Standard 45 columns)
CANONICAL_HEADERS = [
    'LOJA (SOMENTE NÚMEROS)', 'NOME LOJA', 'MATRÍCULA - ID', 'NOME', 'STATUS FUNCIONÁRIO',
    'DT DESLIGAMENTO', 'LÍDER DIRETO', 'CPF (SOMENTE NÚMEROS)', 'DATA DE ADMISSÃO', 'DATA CARGO',
    'RAÇA/ETNIA', 'NACIONALIDADE', 'INSTRUÇÃO DE ENSINO', 'PCD', 'CARGO',
    'DESCRIÇÃO ÁREA', 'DATA DE NASCIMENTO', 'SEXO BIOLÓGICO', 'ESTADO CIVIL', 'E-MAIL',
    'O COLABORADOR É MÃE OU PAI?',
    'NOME DO FILHO 1', 'SEXO DO FILHO 1', 'DATA DE NASCIMENTO DO FILHO 1',
    'NOME DO FILHO 2', 'SEXO DO FILHO 2', 'DATA DE NASCIMENTO DO FILHO 2',
    'NOME DO FILHO 3', 'SEXO DO FILHO 3', 'DATA DE NASCIMENTO DO FILHO 3',
    'NOME DO FILHO 4', 'SEXO DO FILHO 4', 'DATA DE NASCIMENTO DO FILHO 4',
    'NOME DO FILHO 5', 'SEXO DO FILHO 5', 'DATA DE NASCIMENTO DO FILHO 5',
    'NOME DO FILHO 6', 'SEXO DO FILHO 6', 'DATA DE NASCIMENTO DO FILHO 6',
    'NOME DO FILHO 7', 'SEXO DO FILHO 7', 'DATA DE NASCIMENTO DO FILHO 7',
    'NOME DO FILHO 8', 'SEXO DO FILHO 8', 'DATA DE NASCIMENTO DO FILHO 8'
]

def col_to_idx(col_str):
    idx = 0
    for char in col_str:
        if 'A' <= char <= 'Z':
            idx = idx * 26 + (ord(char) - ord('A') + 1)
    return idx - 1

def format_excel_date(val):
    if not val:
        return ''
    val_str = str(val).strip()
    if not val_str or val_str == 'None':
        return ''
    try:
        num = float(val_str)
        if 10000 <= num <= 90000:
            dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=num)
            return dt.strftime('%d/%m/%Y')
    except:
        pass

    parts = re.split(r'[/\.-]', val_str)
    if len(parts) == 2 and len(parts[1]) == 6:
        parts = [parts[0], parts[1][:2], parts[1][2:]]

    if len(parts) == 3:
        p0, p1, p2 = parts[0].strip(), parts[1].strip(), parts[2].strip()
        if len(p0) == 4:
            return f"{p2.zfill(2)}/{p1.zfill(2)}/{p0}"
        elif len(p2) == 4:
            return f"{p0.zfill(2)}/{p1.zfill(2)}/{p2}"
        elif len(p2) == 2:
            full_year = f"20{p2}" if int(p2) < 50 else f"19{p2}"
            return f"{p0.zfill(2)}/{p1.zfill(2)}/{full_year}"

    return val_str

def calculate_exact_age(dob_str, ref_date=None):
    if not dob_str:
        return None
    if ref_date is None:
        ref_date = datetime.date.today()
    
    formatted = format_excel_date(dob_str)
    if not formatted:
        return None
    
    try:
        parts = formatted.split('/')
        if len(parts) == 3:
            day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
            dob = datetime.date(year, month, day)
            
            age = ref_date.year - dob.year
            if (ref_date.month, ref_date.day) < (dob.month, dob.day):
                age -= 1
            return age if age >= 0 else None
    except Exception:
        pass
    return None

def clean_region_name(num, parsed_name=None):
    num_str = str(num).strip()
    if parsed_name:
        clean = re.sub(r'(?i)RELAÇÃO\s*DE\s*FUNCIONÁRIOS\s*PARCEIROS|RELAÇÃO|FUNCIONÁRIOS|PARCEIROS|LOJA\s*\d*|UNIDADE\s*\d*|\.xlsx', '', parsed_name).strip(' -_')
        if clean and len(clean) >= 3 and 'ATUAL' not in clean:
            return clean.upper()
    return STORE_REGIONS_MAP.get(num_str, f"REGIÃO {num_str}")

def extract_store_num_and_name(file_path):
    filename = os.path.basename(file_path)
    parent_dir = os.path.basename(os.path.dirname(file_path))

    if '2810' in filename or '2810' in parent_dir:
        return '2810 CD', 'JOÃO PESSOA (CD)'

    match = re.search(r'LOJA\s*(\d{3,4})', filename, re.IGNORECASE) or re.search(r'(\d{3,4})', parent_dir)
    if match:
        num = match.group(1).strip()
        region = clean_region_name(num)
        return num, region

    return '2050', 'BRASÍLIA'

def parse_xlsx_fast_exact(file_path):
    employees = []
    loja_num, loja_nome = extract_store_num_and_name(file_path)
    store_info = {'loja_num': loja_num, 'nome_loja': loja_nome, 'filename': os.path.basename(file_path)}

    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            shared_strings = []
            if 'xl/sharedStrings.xml' in z.namelist():
                tree = ET.fromstring(z.read('xl/sharedStrings.xml'))
                for si in tree.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                    t_el = si.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                    if t_el is not None and t_el.text:
                        shared_strings.append(t_el.text)
                    else:
                        full_txt = "".join([t.text for t in si.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t') if t.text])
                        shared_strings.append(full_txt)

            sheet_xml = z.read('xl/worksheets/sheet1.xml')
            tree = ET.fromstring(sheet_xml)

            rows = tree.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row')
            if not rows:
                return store_info, []

            data_rows = []
            for row_el in rows:
                cells = row_el.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c')
                row_cells = {}
                for cell in cells:
                    r_ref = cell.attrib.get('r', '')
                    col_str = ''.join([c for c in r_ref if c.isalpha()])
                    if not col_str:
                        continue
                    col_idx = col_to_idx(col_str)

                    cell_type = cell.attrib.get('t', '')
                    val_el = cell.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                    val = val_el.text if val_el is not None else ''

                    if cell_type == 's' and val.isdigit():
                        idx = int(val)
                        if idx < len(shared_strings):
                            val = shared_strings[idx]

                    row_cells[col_idx] = val.strip() if val else ''
                
                if row_cells:
                    max_c = max(row_cells.keys())
                    row_arr = [row_cells.get(i, '') for i in range(max_c + 1)]
                    data_rows.append(row_arr)

            if not data_rows:
                return store_info, []

            header_row_idx = 0
            for idx, r in enumerate(data_rows[:5]):
                r_str = " ".join([str(x).upper() for x in r])
                if 'NOME' in r_str and ('MATRÍCULA' in r_str or 'MATRICULA' in r_str or 'CPF' in r_str or 'STATUS' in r_str):
                    header_row_idx = idx
                    break

            headers = [str(c).strip().upper() for c in data_rows[header_row_idx]]

            def find_col(candidates):
                # 1st pass: exact match to avoid substring false-positives (e.g., 'NOME' matching 'NOME LOJA')
                for cand in candidates:
                    cand_u = cand.upper()
                    for idx, h in enumerate(headers):
                        if cand_u == h.strip():
                            return idx
                # 2nd pass: substring match
                for cand in candidates:
                    cand_u = cand.upper()
                    for idx, h in enumerate(headers):
                        if cand_u in h:
                            return idx
                return -1

            c_loja_num = find_col(['LOJA (SOMENTE NÚMEROS)', 'LOJA SOMENTE'])
            c_loja_nome = find_col(['NOME LOJA'])
            c_matricula = find_col(['MATRÍCULA', 'MATRICULA', 'ID'])
            c_nome = find_col(['NOME'])
            c_status = find_col(['STATUS FUNCIONÁRIO', 'STATUS FUNCIONARIO', 'STATUS'])
            c_dt_deslig = find_col(['DT DESLIGAMENTO', 'DESLIGAMENTO'])
            c_lider = find_col(['LÍDER DIRETO', 'LIDER DIRETO'])
            c_cpf = find_col(['CPF (SOMENTE NÚMEROS)', 'CPF'])
            c_adm = find_col(['DATA DE ADMISSÃO', 'DATA ADMISSAO', 'ADMISSÃO', 'ADMISSAO'])
            c_dt_cargo = find_col(['DATA CARGO'])
            c_raca = find_col(['RAÇA/ETNIA', 'RACA/ETNIA', 'RAÇA', 'ETNIA'])
            c_nacionalidade = find_col(['NACIONALIDADE'])
            c_instrucao = find_col(['INSTRUÇÃO DE ENSINO', 'INSTRUCAO'])
            c_pcd = find_col(['PCD'])
            c_cargo = find_col(['CARGO'])
            c_area = find_col(['DESCRIÇÃO ÁREA', 'DESCRICAO AREA', 'ÁREA', 'AREA'])
            c_dt_nasc = find_col(['DATA DE NASCIMENTO', 'DATA NASCIMENTO'])
            c_sexo = find_col(['SEXO BIOLÓGICO', 'SEXO BIOLOGICO', 'SEXO'])
            c_estado_civil = find_col(['ESTADO CIVIL'])
            c_email = find_col(['E-MAIL', 'EMAIL'])
            c_mae_pai = find_col(['O COLABORADOR É MÃE OU PAI?', 'MÃE OU PAI', 'MAE OU PAI'])

            children_cols = []
            for i in range(1, 9):
                c_fnome = find_col([f'NOME DO FILHO {i}'])
                c_fsexo = find_col([f'SEXO DO FILHO {i}'])
                c_fnasc = find_col([f'DATA DE NASCIMENTO DO FILHO {i}', f'NASCIMENTO DO FILHO {i}'])
                children_cols.append((c_fnome, c_fsexo, c_fnasc))

            ref_today = datetime.date.today()

            for row_idx in range(header_row_idx + 1, len(data_rows)):
                r = data_rows[row_idx]
                if not r:
                    continue
                
                nome = r[c_nome].strip() if c_nome != -1 and c_nome < len(r) else ''
                if not nome or nome.upper() == 'NOME' or 'TOTAL' in nome.upper():
                    continue

                status = r[c_status].strip().upper() if c_status != -1 and c_status < len(r) else 'ATIVO'
                if not status:
                    status = 'ATIVO'

                l_num = store_info['loja_num']
                if c_loja_num != -1 and c_loja_num < len(r) and r[c_loja_num].strip() and r[c_loja_num].strip() != '??':
                    l_num = r[c_loja_num].strip()

                l_nome = clean_region_name(l_num, r[c_loja_nome].strip() if c_loja_nome != -1 and c_loja_nome < len(r) else store_info['nome_loja'])

                dt_adm_raw = r[c_adm].strip() if c_adm != -1 and c_adm < len(r) else ''
                dt_adm = format_excel_date(dt_adm_raw)

                adm_ano, adm_mes, adm_dia = None, None, None
                if dt_adm:
                    parts = dt_adm.split('/')
                    if len(parts) == 3:
                        adm_dia = parts[0]
                        adm_mes = parts[1]
                        adm_ano = parts[2]

                dt_nasc_raw = r[c_dt_nasc].strip() if c_dt_nasc != -1 and c_dt_nasc < len(r) else ''
                dt_nasc = format_excel_date(dt_nasc_raw)
                emp_idade = calculate_exact_age(dt_nasc, ref_today)

                eh_mae_pai = r[c_mae_pai].strip().upper() if c_mae_pai != -1 and c_mae_pai < len(r) else 'NÃO'
                if 'SIM' in eh_mae_pai: eh_mae_pai = 'SIM'
                else: eh_mae_pai = 'NÃO'

                filhos = []
                for fnome_i, fsexo_i, fnasc_i in children_cols:
                    fnome = r[fnome_i].strip() if fnome_i != -1 and fnome_i < len(r) else ''
                    fsexo = r[fsexo_i].strip() if fsexo_i != -1 and fsexo_i < len(r) else ''
                    fnasc_raw = r[fnasc_i].strip() if fnasc_i != -1 and fnasc_i < len(r) else ''
                    fnasc = format_excel_date(fnasc_raw)

                    if fnome or fnasc:
                        f_idade = calculate_exact_age(fnasc, ref_today)
                        filhos.append({
                            'nome': fnome or 'Filho Sem Nome',
                            'sexo': fsexo,
                            'data_nascimento': fnasc,
                            'idade': f_idade
                        })

                if len(filhos) > 0:
                    eh_mae_pai = 'SIM'

                clean_nome_part = re.sub(r'\W+', '', nome)[:10]
                emp_record = {
                    'id': f"{l_num}_{row_idx}_{clean_nome_part}",
                    'loja_num': l_num,
                    'nome_loja': l_nome,
                    'matricula': r[c_matricula].strip() if c_matricula != -1 and c_matricula < len(r) else '',
                    'nome': nome,
                    'status': status,
                    'dt_desligamento': format_excel_date(r[c_dt_deslig]) if c_dt_deslig != -1 and c_dt_deslig < len(r) else '',
                    'lider_direto': r[c_lider].strip() if c_lider != -1 and c_lider < len(r) else '',
                    'cpf': r[c_cpf].strip() if c_cpf != -1 and c_cpf < len(r) else '',
                    'dt_admissao': dt_adm,
                    'adm_ano': adm_ano,
                    'adm_mes': adm_mes,
                    'adm_dia': adm_dia,
                    'raca_etnia': r[c_raca].strip() if c_raca != -1 and c_raca < len(r) else 'NÃO INFORMADO',
                    'instrucao': r[c_instrucao].strip() if c_instrucao != -1 and c_instrucao < len(r) else '',
                    'pcd': 'SIM' if c_pcd != -1 and c_pcd < len(r) and 'SIM' in r[c_pcd].strip().upper() else 'NÃO',
                    'cargo': r[c_cargo].strip() if c_cargo != -1 and c_cargo < len(r) else '',
                    'area': r[c_area].strip() if c_area != -1 and c_area < len(r) else '',
                    'dt_nascimento': dt_nasc,
                    'idade': emp_idade,
                    'sexo': r[c_sexo].strip().upper() if c_sexo != -1 and c_sexo < len(r) else '',
                    'email': r[c_email].strip() if c_email != -1 and c_email < len(r) else '',
                    'eh_mae_pai': eh_mae_pai,
                    'qtd_filhos': len(filhos),
                    'filhos': filhos
                }
                employees.append(emp_record)

    except Exception as e:
        print(f"Error parsing {file_path}: {e}", flush=True)

    return store_info, employees

def scan_and_rebuild_dataset(force=False):
    if not os.path.exists(PARTNERS_DIR):
        print(f"Partners directory not found: {PARTNERS_DIR}", flush=True)
        return

    files = []
    for root, dirs, f_names in os.walk(PARTNERS_DIR):
        for f in f_names:
            if f.endswith('.xlsx') and not f.startswith('~$'):
                files.append(os.path.join(root, f))
    
    file_mtimes = {f: os.path.getmtime(f) for f in files}

    if not force:
        with global_lock:
            if file_mtimes == global_state['file_mtimes']:
                return

    print(f"🔄 Escaneando {len(files)} planilhas Excel em {PARTNERS_DIR}...", flush=True)

    all_employees = []
    stores_summary_parsed = []
    admission_years_set = set()

    for f_path in sorted(files):
        s_info, emps = parse_xlsx_fast_exact(f_path)
        if emps:
            actives = sum(1 for e in emps if 'ATIVO' in e['status'])
            terms = sum(1 for e in emps if 'DESLIG' in e['status'])
            s_info['total_colaboradores'] = len(emps)
            s_info['ativos'] = actives
            s_info['desligados'] = terms
            stores_summary_parsed.append(s_info)

            for e in emps:
                all_employees.append(e)
                if e.get('adm_ano'):
                    admission_years_set.add(e['adm_ano'])

    # Append custom registered employees from custom_employees.json
    custom_emps = load_custom_employees()
    if custom_emps:
        ref_today = datetime.date.today()
        for ce in custom_emps:
            for f in ce.get('filhos', []):
                if f.get('data_nascimento'):
                    f['idade'] = calculate_exact_age(f['data_nascimento'], ref_today)
            all_employees.append(ce)
            if ce.get('adm_ano'):
                admission_years_set.add(ce['adm_ano'])

    # Merge parsed stores with PREDEFINED_STORES master list
    parsed_stores_map = {s['loja_num']: s for s in stores_summary_parsed}
    final_stores_list = []

    for p_store in PREDEFINED_STORES:
        num = p_store['loja_num']
        region = clean_region_name(num)
        if num in parsed_stores_map:
            parsed = parsed_stores_map[num]
            final_stores_list.append({
                'loja_num': num,
                'nome_loja': region,
                'filename': parsed.get('filename', f'LOJA {num}.xlsx'),
                'total_colaboradores': parsed.get('total_colaboradores', 0),
                'ativos': parsed.get('ativos', 0),
                'desligados': parsed.get('desligados', 0)
            })
        else:
            final_stores_list.append({
                'loja_num': num,
                'nome_loja': region,
                'filename': 'Pré-cadastrada',
                'total_colaboradores': 0,
                'ativos': 0,
                'desligados': 0
            })

    total_records = len(all_employees)
    active_records = sum(1 for e in all_employees if 'ATIVO' in e['status'])
    terminated_records = sum(1 for e in all_employees if 'DESLIG' in e['status'])
    mothers_fathers_count = sum(1 for e in all_employees if e['eh_mae_pai'] == 'SIM')
    total_children_count = sum(e['qtd_filhos'] for e in all_employees)
    pcd_count = sum(1 for e in all_employees if e['pcd'] == 'SIM')

    children_by_age = {}
    for e in all_employees:
        for f in e['filhos']:
            if f['idade'] is not None:
                age_key = str(f['idade'])
                children_by_age[age_key] = children_by_age.get(age_key, 0) + 1

    with global_lock:
        global_state['last_updated'] = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        global_state['today_date'] = datetime.date.today().strftime('%d/%m/%Y')
        global_state['current_year'] = datetime.date.today().year
        global_state['stores_count'] = len(final_stores_list)
        global_state['total_records'] = total_records
        global_state['active_records'] = active_records
        global_state['terminated_records'] = terminated_records
        global_state['mothers_fathers_count'] = mothers_fathers_count
        global_state['total_children_count'] = total_children_count
        global_state['children_by_age'] = children_by_age
        global_state['admission_years'] = sorted(list(admission_years_set), reverse=True)
        global_state['pcd_count'] = pcd_count
        global_state['stores'] = final_stores_list
        global_state['employees'] = all_employees
        global_state['file_mtimes'] = file_mtimes

    save_cache_file()
    notify_subscribers()
    print(f"✅ Base reconstruída com sucesso! Total: {total_records} colaboradores em {len(final_stores_list)} lojas limpas (Ex: Loja 2050 - BRASÍLIA).", flush=True)

def save_cache_file():
    try:
        with global_lock:
            cache_data = {
                'last_updated': global_state['last_updated'],
                'today_date': global_state['today_date'],
                'current_year': global_state['current_year'],
                'stores_count': global_state['stores_count'],
                'total_records': global_state['total_records'],
                'active_records': global_state['active_records'],
                'terminated_records': global_state['terminated_records'],
                'mothers_fathers_count': global_state['mothers_fathers_count'],
                'total_children_count': global_state['total_children_count'],
                'children_by_age': global_state['children_by_age'],
                'admission_years': global_state['admission_years'],
                'pcd_count': global_state['pcd_count'],
                'stores': global_state['stores'],
                'employees': global_state['employees']
            }
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving cache file: {e}", flush=True)

def generate_consolidated_csv(employees):
    try:
        import csv
        with open(CONSOLIDATED_CSV, 'w', encoding='utf-8-sig', newline='') as f:
            f.write("sep=;\n")
            writer = csv.writer(f, delimiter=';')
            writer.writerow(CANONICAL_HEADERS)

            for e in employees:
                row = [
                    e.get('loja_num', ''),
                    e.get('nome_loja', ''),
                    e.get('matricula', ''),
                    e.get('nome', ''),
                    e.get('status', ''),
                    e.get('dt_desligamento', ''),
                    e.get('lider_direto', ''),
                    e.get('cpf', ''),
                    e.get('dt_admissao', ''),
                    '',
                    e.get('raca_etnia', ''),
                    'BRASILEIRA',
                    e.get('instrucao', ''),
                    e.get('pcd', ''),
                    e.get('cargo', ''),
                    e.get('area', ''),
                    e.get('dt_nascimento', ''),
                    e.get('sexo', ''),
                    '',
                    e.get('email', ''),
                    e.get('eh_mae_pai', '')
                ]
                filhos = e.get('filhos', [])
                for i in range(8):
                    if i < len(filhos):
                        f_item = filhos[i]
                        row.extend([f_item.get('nome', ''), f_item.get('sexo', ''), f_item.get('data_nascimento', '')])
                    else:
                        row.extend(['', '', ''])
                writer.writerow(row)
    except Exception as ex:
        print(f"Error generating CSV: {ex}", flush=True)

def generate_consolidated_excel(employees):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consolidado Parceiros"

        ws.views.sheetView[0].showGridLines = True

        header_fill = PatternFill(start_color="141820", end_color="141820", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="C0F200")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        ws.append(CANONICAL_HEADERS)
        for col_num in range(1, len(CANONICAL_HEADERS) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[1].height = 28

        for e in employees:
            row = [
                e.get('loja_num', ''),
                e.get('nome_loja', ''),
                e.get('matricula', ''),
                e.get('nome', ''),
                e.get('status', ''),
                e.get('dt_desligamento', ''),
                e.get('lider_direto', ''),
                e.get('cpf', ''),
                e.get('dt_admissao', ''),
                '',
                e.get('raca_etnia', ''),
                'BRASILEIRA',
                e.get('instrucao', ''),
                e.get('pcd', ''),
                e.get('cargo', ''),
                e.get('area', ''),
                e.get('dt_nascimento', ''),
                e.get('sexo', ''),
                '',
                e.get('email', ''),
                e.get('eh_mae_pai', '')
            ]
            filhos = e.get('filhos', [])
            for i in range(8):
                if i < len(filhos):
                    f_item = filhos[i]
                    row.extend([f_item.get('nome', ''), f_item.get('sexo', ''), f_item.get('data_nascimento', '')])
                else:
                    row.extend(['', '', ''])
            ws.append(row)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(CANONICAL_HEADERS)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(min(max_len + 3, 40), 12)

        wb.save(CONSOLIDATED_XLSX)
        return CONSOLIDATED_XLSX
    except Exception as ex:
        print(f"Error generating XLSX: {ex}", flush=True)
        return None

def notify_subscribers():
    msg = f"data: {json.dumps({'type': 'update', 'last_updated': global_state['last_updated']})}\n\n"
    with global_lock:
        subscribers = list(global_state['subscribers'])
    for sub in subscribers:
        try:
            sub.write(msg.encode('utf-8'))
            sub.flush()
        except Exception:
            pass

def background_watcher():
    while True:
        try:
            scan_and_rebuild_dataset()
        except Exception as e:
            print(f"Watcher error: {e}", flush=True)
        time.sleep(3)

class DashboardRequestHandler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path

        if path == '/' or path == '/index.html':
            self.serve_static_file(os.path.join(CURRENT_DIR, 'index.html'), 'text/html; charset=utf-8')
        elif path == '/leo_logo.png':
            self.serve_static_file(os.path.join(CURRENT_DIR, 'leo_logo.png'), 'image/png')
        elif path == '/api/data':
            self.send_json_response(global_state)
        elif path == '/api/events':
            self.serve_sse_stream()
        elif path == '/api/export':
            self.serve_export_download()
        elif path == '/api/users':
            users = load_users()
            clean_users = [{'id': u['id'], 'nome': u['nome'], 'email': u['email'], 'role': u['role'], 'stores': u['stores']} for u in users]
            self.send_json_response_direct(clean_users)
        else:
            self.send_error(404, "Not Found")

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path

        content_length = int(self.headers.get('Content-Length', 0))
        body_bytes = self.rfile.read(content_length) if content_length > 0 else b'{}'
        
        try:
            payload = json.loads(body_bytes.decode('utf-8'))
        except Exception:
            payload = {}

        if path == '/api/login':
            email = payload.get('email', '').strip().lower()
            senha = payload.get('senha', '').strip()
            users = load_users()
            matched = next((u for u in users if u['email'].lower() == email and u['senha'] == senha), None)
            if matched:
                res_user = {
                    'id': matched['id'],
                    'nome': matched['nome'],
                    'email': matched['email'],
                    'role': matched['role'],
                    'stores': matched['stores']
                }
                self.send_json_response_direct({'success': True, 'user': res_user})
            else:
                self.send_json_response_direct({'success': False, 'message': 'E-mail ou senha incorretos.'}, status=401)

        elif path == '/api/users/save':
            users = load_users()
            u_id = payload.get('id')
            nome = payload.get('nome', '').strip()
            email = payload.get('email', '').strip()
            senha = payload.get('senha', '').strip()
            role = payload.get('role', 'USER')
            stores = payload.get('stores', [])

            if not nome or not email:
                self.send_json_response_direct({'success': False, 'message': 'Nome e E-mail são obrigatórios.'}, status=400)
                return

            if u_id:
                for u in users:
                    if u['id'] == u_id:
                        u['nome'] = nome
                        u['email'] = email
                        if senha: u['senha'] = senha
                        u['role'] = role
                        u['stores'] = stores
                        break
            else:
                new_id = f"usr_{int(time.time())}"
                new_user = {
                    'id': new_id,
                    'nome': nome,
                    'email': email,
                    'senha': senha or '123456',
                    'role': role,
                    'stores': stores
                }
                users.append(new_user)

            save_users(users)
            self.send_json_response_direct({'success': True, 'users': users})

        elif path == '/api/users/delete':
            u_id = payload.get('id')
            users = load_users()
            users = [u for u in users if u['id'] != u_id]
            save_users(users)
            self.send_json_response_direct({'success': True, 'users': users})

        elif path == '/api/employees/add':
            loja_num = payload.get('loja_num', '').strip()
            nome = payload.get('nome', '').strip()

            if not loja_num or not nome:
                self.send_json_response_direct({'success': False, 'message': 'Loja e Nome são obrigatórios.'}, status=400)
                return

            ref_today = datetime.date.today()
            dt_adm = format_excel_date(payload.get('dt_admissao', ''))
            adm_ano, adm_mes, adm_dia = None, None, None
            if dt_adm:
                parts = dt_adm.split('/')
                if len(parts) == 3:
                    adm_dia, adm_mes, adm_ano = parts[0], parts[1], parts[2]

            dt_nasc = format_excel_date(payload.get('dt_nascimento', ''))
            emp_idade = calculate_exact_age(dt_nasc, ref_today)

            raw_filhos = payload.get('filhos', [])
            filhos = []
            for f in raw_filhos:
                fn = f.get('nome', '').strip()
                fs = f.get('sexo', '').strip().upper()
                fd = format_excel_date(f.get('data_nascimento', ''))
                fi = calculate_exact_age(fd, ref_today)
                if fn or fd:
                    filhos.append({
                        'nome': fn or 'Filho Sem Nome',
                        'sexo': fs,
                        'data_nascimento': fd,
                        'idade': fi
                    })

            eh_mae_pai = 'SIM' if len(filhos) > 0 or payload.get('eh_mae_pai') == 'SIM' else 'NÃO'

            clean_nome_part = re.sub(r'\W+', '', nome)[:10]
            new_emp = {
                'id': f"custom_{int(time.time())}_{clean_nome_part}",
                'loja_num': loja_num,
                'nome_loja': clean_region_name(loja_num, payload.get('nome_loja')),
                'matricula': payload.get('matricula', '').strip(),
                'nome': nome,
                'status': payload.get('status', 'ATIVO').strip().upper(),
                'dt_desligamento': format_excel_date(payload.get('dt_desligamento', '')),
                'lider_direto': payload.get('lider_direto', '').strip(),
                'cpf': payload.get('cpf', '').strip(),
                'dt_admissao': dt_adm,
                'adm_ano': adm_ano,
                'adm_mes': adm_mes,
                'adm_dia': adm_dia,
                'raca_etnia': payload.get('raca_etnia', 'NÃO INFORMADO').strip().upper(),
                'instrucao': payload.get('instrucao', '').strip(),
                'pcd': payload.get('pcd', 'NÃO').strip().upper(),
                'cargo': payload.get('cargo', '').strip(),
                'area': payload.get('area', '').strip(),
                'dt_nascimento': dt_nasc,
                'idade': emp_idade,
                'sexo': payload.get('sexo', '').strip().upper(),
                'email': payload.get('email', '').strip(),
                'eh_mae_pai': eh_mae_pai,
                'qtd_filhos': len(filhos),
                'filhos': filhos
            }

            custom_list = load_custom_employees()
            custom_list.append(new_emp)
            save_custom_employees(custom_list)

            # Rebuild dataset
            scan_and_rebuild_dataset(force=True)
            self.send_json_response_direct({'success': True, 'employee': new_emp})

        else:
            self.send_error(404, "Not Found")

    def serve_static_file(self, filepath, content_type):
        if not os.path.exists(filepath):
            self.send_error(404, "File Not Found")
            return
        with open(filepath, 'rb') as f:
            content = f.read()
        self.send_response(200)
        self.send_header('Content-Type', content_type)
        self.send_header('Content-Length', str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def send_json_response_direct(self, obj, status=200):
        body = json.dumps(obj, ensure_ascii=False).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_json_response(self, data):
        with global_lock:
            clean_data = {
                'last_updated': data['last_updated'],
                'today_date': data['today_date'],
                'current_year': data['current_year'],
                'stores_count': data['stores_count'],
                'total_records': data['total_records'],
                'active_records': data['active_records'],
                'terminated_records': data['terminated_records'],
                'mothers_fathers_count': data['mothers_fathers_count'],
                'total_children_count': data['total_children_count'],
                'children_by_age': data['children_by_age'],
                'admission_years': data['admission_years'],
                'pcd_count': data['pcd_count'],
                'stores': data['stores'],
                'cargos': sorted(list(set(e['cargo'] for e in data.get('employees', []) if e.get('cargo')))),
                'employees': data['employees']
            }
            body = json.dumps(clean_data, ensure_ascii=False).encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def serve_sse_stream(self):
        self.send_response(200)
        self.send_header('Content-Type', 'text/event-stream')
        self.send_header('Cache-Control', 'no-cache')
        self.send_header('Connection', 'keep-alive')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()

        sub_wfile = self.wfile
        with global_lock:
            global_state['subscribers'].append(sub_wfile)

        try:
            init_msg = f"data: {json.dumps({'type': 'connected', 'last_updated': global_state['last_updated']})}\n\n"
            sub_wfile.write(init_msg.encode('utf-8'))
            sub_wfile.flush()
            while True:
                time.sleep(15)
                ping = f": ping\n\n"
                sub_wfile.write(ping.encode('utf-8'))
                sub_wfile.flush()
        except Exception:
            pass
        finally:
            with global_lock:
                if sub_wfile in global_state['subscribers']:
                    global_state['subscribers'].remove(sub_wfile)

    def serve_export_download(self):
        xlsx_path = generate_consolidated_excel(global_state.get('employees', []))
        if xlsx_path and os.path.exists(xlsx_path):
            with open(xlsx_path, 'rb') as f:
                content = f.read()
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="CONSOLIDADO_PARCEIROS.xlsx"')
            self.send_header('Content-Length', str(len(content)))
            self.end_headers()
            self.wfile.write(content)
        else:
            generate_consolidated_csv(global_state.get('employees', []))
            with open(CONSOLIDATED_CSV, 'rb') as f:
                content = f.read()
            self.send_response(200)
            self.send_header('Content-Type', 'text/csv; charset=utf-8-sig')
            self.send_header('Content-Disposition', 'attachment; filename="CONSOLIDADO_PARCEIROS.csv"')
            self.send_header('Content-Length', str(len(content)))
            self.end_headers()
            self.wfile.write(content)

class ThreadedHTTPServer(ThreadingMixIn, HTTPServer):
    """Handle requests in a separate thread."""
    daemon_threads = True

def run_server(host='127.0.0.1', port=3000):
    print("==========================================================", flush=True)
    print("  PARCEIROS LEO - DASHBOARD UNIFICADO (STORE REGIONS CLEAN)", flush=True)
    print("==========================================================", flush=True)
    print(f"OneDrive path: {PARTNERS_DIR}", flush=True)

    scan_and_rebuild_dataset(force=True)

    watcher_thread = threading.Thread(target=background_watcher, daemon=True)
    watcher_thread.start()

    server_address = (host, port)
    httpd = ThreadedHTTPServer(server_address, DashboardRequestHandler)
    print(f"\n🚀 Parceiros Leo rodando em: http://{host}:{port}", flush=True)
    print("⚡ Pressione Ctrl+C para encerrar o servidor.\n", flush=True)
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nServidor encerrado.", flush=True)

if __name__ == '__main__':
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 3000
    run_server('127.0.0.1', port)
